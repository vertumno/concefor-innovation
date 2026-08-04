# -*- coding: utf-8 -*-
"""Gera a planilha ATUAÇÃO DA CGTE NO CONCEFOR — VIII Concefor (17-20/08/2026).
Estrutura espelhada na planilha de 2024 (uma aba por dia, 7 colunas).

⚠️ ATENÇÃO: rodar este script SOBRESCREVE o .xlsx da pasta acima. Se a planilha já
foi preenchida com os nomes da equipe, edite a planilha direto — não regere daqui.
O script serve para o caso de a programação mudar e valer a pena refazer do zero.

Uso:  python gerar-planilha-atuacao-2026.py       (requer openpyxl)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY900 = "FF102A5C"
NAVY800 = "FF173B79"
TEAL = "FF0E8FA8"
GOLD = "FFF5B82E"
PAPER = "FFF4F7FB"
LINE = "FFD9E2EE"
INK = "FF0E2240"
DESTAQUE = "FFE8F4F8"   # linhas com transmissão + Libras
CULTURAL = "FFFFF6E0"   # momentos de coffee/cultural

COLS = [
    ("PROGRAMAÇÃO", 52),
    ("LOCAL", 22),
    ("TRANSMISSÃO E FILMAGEM", 26),
    ("LIBRAS", 24),
    ("PODCAST\n(responsável pelos termos: ___)", 26),
    ("COBERTURA REELS/STORIES E FOTOGRAFIA\n(responsável pelas postagens: ___)", 30),
    ("OBSERVAÇÕES", 46),
]

thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TITULO = "ATUAÇÃO DA CGTE NO CONCEFOR"
SUBTITULO = ("VIII CONCEFOR · 17 a 20 de agosto de 2026 · Cefor/Ifes · Vitória-ES · "
             "20 anos do Cefor")

# ---------------------------------------------------------------- conteúdo ---
# (programação, local, transmissão, libras, podcast, cobertura, observações, tipo)
# tipo: 'destaque' (palestra/mesa com transmissão), 'cultural', None

DIAS = [
    {
        "aba": "1708 - SEGUNDA",
        "dia": "DIA: 17/08/26 (SEGUNDA-FEIRA) — ABERTURA",
        "linhas": [
            ("13h30  Credenciamento",
             "🔴 a definir", "", "", "",
             "___",
             "Chegada da equipe e montagem: definir horário de call da CGTE. "
             "Primeiro dia de cobertura — checar carga de baterias, cartões e etiquetas dos crachás.",
             None),
            ("14h30 às 17h30  Eventos UAB / UnAC / NTE",
             "🟡 UAB e UnAC: Pátio\nNTE: Auditório\n(a confirmar)",
             "", "", "", "___",
             "Locais ainda a confirmar com a Coordenação (pedido em 28/07).",
             None),
            ("18h00  Momento Musical",
             "Pátio", "", "", "", "___",
             "", "cultural"),
            ("18h30  Abertura do Evento",
             "Pátio", "Equipe audiovisual: ___", "Equipe Libras: ___", "", "___",
             "🔒 TRANSMISSÃO NO AR das 18h30 às 20h30 — nenhum podcast neste intervalo.",
             "destaque"),
            ("19h00 às 20h00  Palestra “20 Anos de EaD e o Cefor” – Dra. Vanessa Battestin",
             "Pátio", "Equipe audiovisual: ___", "Equipe Libras: ___",
             "Não grava hoje.\nPodcast dela: 18/08 8h",
             "___",
             "Quem recebe a palestrante: ___\nTermo de imagem/voz assinado ANTES da gravação do "
             "podcast (modelo ainda a providenciar).",
             "destaque"),
            ("20h00 às 20h30  Lançamentos (e-book 20 anos · e-book do NTE · plataforma de MOOCs · "
             "nova Base de Conhecimento)",
             "Pátio", "Equipe audiovisual: ___", "Equipe Libras: ___", "", "___",
             "São 4 lançamentos em sequência — garantir foto/registro de cada um.",
             "destaque"),
            ("20h30  Celebração do aniversário de 20 anos do Cefor (bolo)",
             "Pátio", "", "Equipe Libras: ___", "", "___",
             "⭐ Momento mais forte do evento para reels. Bolo com a arte da Andreia.",
             "cultural"),
        ],
    },
    {
        "aba": "1808 - TERÇA",
        "dia": "DIA: 18/08/26 (TERÇA-FEIRA)",
        "linhas": [
            ("08h00  Mostra de Produtos e Produções Técnicas · Painel",
             "🔴 a definir", "", "",
             "🎙 PODCAST 1 — Vanessa Battestin\n8h00 às 9h00\n(alt.: almoço 12h–13h30)",
             "___",
             "Pauta: 20 anos de EaD e o Cefor. Host definido pela Coordenação (🔴 pendente).",
             None),
            ("09h30  Mesa-Redonda “Tecnologia Delas” – Dra. Márcia Oliveira; Dra. Jaqueline Sanz; "
             "Dra. Mariella Berger",
             "Pátio", "Equipe audiovisual: ___", "Equipe Libras: ___",
             "APÓS A MESA\n(gravação no almoço)", "___",
             "🔒 Transmissão no ar 9h30–11h30. Quem recebe as palestrantes: ___",
             "destaque"),
            ("11h30  Mostra de Produtos e Produções Técnicas · Painel",
             "🔴 a definir", "", "", "", "___",
             "", None),
            ("12h00 às 13h30  Almoço",
             "—", "", "",
             "🎙 PODCAST 2 — Mesa “Tecnologia Delas” (1 a 2 participantes)\n12h00 às 13h30",
             "",
             "Alternativa: 13h30–14h15, durante as sessões técnicas.",
             None),
            ("13h30  Sessões técnicas · apresentação de artigos completos",
             "🔴 a definir\n(salas 1 a 4?)", "", "", "", "___",
             "Apoio e organização ao longo do dia: ___",
             None),
            ("14h30  Palestra internacional – Felipe Maciel Tessarolo (The Open University, UK)",
             "Pátio", "Equipe audiovisual: ___", "Equipe Libras: ___",
             "APÓS A PALESTRA\n(16h20)", "___",
             "🔒 Transmissão no ar 14h30–16h. Quem recebe o palestrante: ___",
             "destaque"),
            ("16h00  Coffee-break",
             "Pátio", "", "", "", "___",
             "", "cultural"),
            ("16h20 às 17h00  (pós-coffee)",
             "—", "", "",
             "🎙 PODCAST 3 — Felipe Tessarolo\n16h20 às 17h00\n(alt.: 19/08 8h–9h)",
             "",
             "Pauta: IA, Open University, acessibilidade.",
             None),
        ],
    },
    {
        "aba": "1908 - QUARTA",
        "dia": "DIA: 19/08/26 (QUARTA-FEIRA)",
        "linhas": [
            ("08h00  Sessões técnicas · apresentação de artigos completos",
             "🔴 a definir\n(salas 1 a 4?)", "", "", "", "___",
             "Janela alternativa do podcast do Tessarolo, se não tiver rolado ontem.",
             None),
            ("09h00  Momento Cultural",
             "Pátio", "", "", "", "___",
             "", "cultural"),
            ("09h30  Mesa-Redonda “Desafios da EaD para os próximos 20 anos” – "
             "Felipe Maciel Tessarolo; Dra. Rutinelli da Penha Fávero",
             "Pátio", "Equipe audiovisual: ___", "Equipe Libras: ___",
             "APÓS A MESA\n(gravação no almoço)", "___",
             "🔒 Transmissão no ar 9h30–11h30.",
             "destaque"),
            ("12h00 às 13h30  Almoço",
             "—", "", "",
             "🎙 PODCAST 4 — Mesa “Desafios da EaD”\n12h00 às 13h30",
             "",
             "Alternativa: 13h30–14h15, durante as sessões técnicas.",
             None),
            ("13h30  Sessões técnicas · apresentação de artigos completos",
             "🔴 a definir\n(salas 1 a 4?)", "", "", "", "___",
             "Apoio e organização ao longo do dia: ___",
             None),
            ("14h30  Palestra “Autoria com IA generativa: desafios para a EaD” – "
             "Dr. Mariano Pimentel",
             "Pátio", "Equipe audiovisual: ___", "Equipe Libras: ___",
             "APÓS A PALESTRA\n⚠️ ver observação", "___",
             "🔒 Transmissão no ar 14h30–16h. Quem recebe o palestrante: ___",
             "destaque"),
            ("16h00  Coffee-break",
             "Pátio", "", "", "", "___",
             "", "cultural"),
            ("16h30  Palestra “Inovações da IA para a Educação” – Dr. Mauro Oliveira",
             "Pátio", "Equipe audiovisual: ___", "Equipe Libras: ___",
             "", "___",
             "⚠️ CONFLITO A RESOLVER: o plano de podcasts marca o episódio do Mariano Pimentel "
             "para 16h20–17h (pós-coffee), mas esta palestra entrou na programação depois (21/07) "
             "e ocupa a mesma janela com transmissão no ar. "
             "Opções: gravar o Mariano no almoço de 20/08 ou logo após esta palestra (a partir das 18h). "
             "Quem recebe o palestrante: ___",
             "destaque"),
        ],
    },
    {
        "aba": "2008 - QUINTA",
        "dia": "DIA: 20/08/26 (QUINTA-FEIRA) — ENCERRAMENTO",
        "linhas": [
            ("09h00 às 12h00  Momento I · Eventos paralelos (6 eventos simultâneos — ver abaixo)",
             "Pátio, Salas 1, 2 e 4,\nAuditório e Laboratório\nde Informática",
             "Definir o que é filmado\n(1 câmera por evento?)",
             "___", "", "___",
             "6 eventos ao mesmo tempo — a cobertura precisa de rodízio entre as salas. "
             "Apoio e organização: ___",
             None),
            ("      · IA além do chat: da célula ao organismo baseado em IA (oficina, máx. 20)",
             "Laboratório de\nInformática", "", "", "", "___",
             "⚠️ Ministrada por MARQUITO e ELTON — os dois ficam FORA da operação de cobertura "
             "o dia inteiro (9h–12h e 13h30–16h). Apoio: ___",
             None),
            ("      · “Ciência delas” no contexto do Projeto Rio Doce Escolar (mesa-redonda)",
             "Auditório", "", "", "", "___",
             "Coordenação: Manuella Villar Amado.", None),
            ("      · Escola de Inovação: 6º ano de popularização de novas tecnologias digitais",
             "Sala 2", "", "", "", "___",
             "Coordenação: Patrícia Piana de Andrade e Daniel Moreira dos Santos (CCEC-EI/SEME/PMV).",
             None),
            ("      · Workshop Pros@tec: Desafios da Educação em Computação e da Informática na Educação",
             "Sala 1", "", "", "", "___",
             "Coordenação: Márcia Gonçalves de Oliveira e Rosane Muñoz.", None),
            ("12h00 às 13h30  Almoço",
             "—", "", "",
             "🎙 PODCAST 5 / repescagem\n12h00 às 13h30", "",
             "Janela para o episódio do Mariano Pimentel (ver conflito de 19/08) ou episódio extra.",
             None),
            ("13h30 às 16h00  Momento II · Eventos paralelos",
             "Pátio, Salas 1, 2 e 4,\nAuditório e Laboratório\nde Informática",
             "Definir o que é filmado", "___", "", "___",
             "Continuam: IA além do chat · Escola de Inovação · Pros@tec. "
             "Entram: Entre Dois Mundos (Sala 4) e Educimat 15 anos (Pátio).",
             None),
            ("      · Entre Dois Mundos: uma aventura para aprender IA (oficina)",
             "Sala 4", "", "", "", "___",
             "Coordenação: Juliana Cristina dos Santos Waichert.", None),
            ("      · EDUCIMAT: 15 anos elaborando produtos, tecendo saberes e modificando vidas",
             "Pátio", "___", "___", "", "___",
             "Celebração dos 15 anos do Educimat — Edmar Reis Thiengo e "
             "Ana Raquel Santos de Medeiros Garcia. É no Pátio, então cai junto com a cobertura principal.",
             None),
            ("16h00  Premiação dos Melhores Trabalhos do VIII Concefor",
             "Pátio", "Equipe audiovisual: ___", "Equipe Libras: ___", "", "___",
             "🔒 Transmissão no ar. ⭐ Foto dos vencedores para a notícia de encerramento.",
             "destaque"),
            ("16h30  Coffee-break e Momento Cultural · Encerramento",
             "Pátio", "", "___", "", "___",
             "⭐ Foto oficial dos servidores e colaboradores do Concefor — definir horário e ponto "
             "de encontro (em 2024 foi no palco).",
             "cultural"),
        ],
    },
]

# ---------------------------------------------------------------- montagem ---
wb = openpyxl.Workbook()
wb.remove(wb.active)


def cabecalho(ws, texto_dia):
    ws.merge_cells("A1:G2")
    c = ws["A1"]
    c.value = TITULO + "\n" + SUBTITULO
    c.font = Font(name="Calibri", size=15, bold=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY900)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:G3")
    c = ws["A3"]
    c.value = texto_dia
    c.font = Font(name="Calibri", size=12, bold=True, color=INK)
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    for i, (titulo, larg) in enumerate(COLS, start=1):
        c = ws.cell(row=4, column=i, value=titulo)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY800)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[4].height = 46


for dia in DIAS:
    ws = wb.create_sheet(dia["aba"])
    cabecalho(ws, dia["dia"])

    r = 5
    for linha in dia["linhas"]:
        *valores, tipo = linha
        fill = None
        if tipo == "destaque":
            fill = PatternFill("solid", fgColor=DESTAQUE)
        elif tipo == "cultural":
            fill = PatternFill("solid", fgColor=CULTURAL)

        for i, v in enumerate(valores, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            c.font = Font(name="Calibri", size=10,
                          bold=(i == 1 and tipo == "destaque"), color=INK)
            if fill:
                c.fill = fill
        ws.row_dimensions[r].height = 54
        r += 1

    ws.freeze_panes = "B5"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

# ------------------------------------------------------------ aba de apoio ---
ws = wb.create_sheet("EQUIPE E LEGENDA", 0)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 62

ws.merge_cells("A1:C2")
c = ws["A1"]
c.value = TITULO + "\n" + SUBTITULO
c.font = Font(name="Calibri", size=15, bold=True, color="FFFFFFFF")
c.fill = PatternFill("solid", fgColor=NAVY900)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 24
ws.row_dimensions[2].height = 20


def secao(ws, row, titulo):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=titulo)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    return row + 1


def linha(ws, row, a, b, c_, bold=False, wrap=True):
    for i, v in enumerate((a, b, c_), start=1):
        cel = ws.cell(row=row, column=i, value=v)
        cel.font = Font(name="Calibri", size=10, bold=bold, color=INK)
        cel.alignment = Alignment(vertical="top", wrap_text=wrap)
        cel.border = BORDER
    return row + 1


r = 4
r = secao(ws, r, "EQUIPE DA CGTE")
r = linha(ws, r, "PESSOA", "PAPEL", "COMPETÊNCIAS / FRENTE NATURAL", bold=True)
for p, papel, comp in [
    ("Marquito (Marcos V. F. Accioly)", "Coordenador", "Design · Programação com IA · app do evento"),
    ("Elton (Elton Vinicius Silva)", "Designer / Dev", "Design · Programação com IA"),
    ("Tiago", "Audiovisual", "Vídeo · Fotografia · transmissão"),
    ("Leonardo", "Audiovisual", "Vídeo · Fotografia · transmissão"),
    ("Nanda", "Estagiária", "Audiovisual · Fotografia"),
    ("Andreia", "Designer", "Design"),
    ("Juliana", "Designer (banners)", "Design · também coordena um evento paralelo em 20/08"),
    ("Raquel", "Apoio", "Organização e logística"),
    ("Carol", "Intérprete", "Libras"),
    ("Eliana", "Intérprete", "Libras"),
]:
    r = linha(ws, r, p, papel, comp)

r += 1
r = secao(ws, r, "REGRAS FIXAS (decisões já tomadas)")
for t, d in [
    ("Transmissão + Libras",
     "Decidido em 16/07/2026: TODAS as palestras e mesas-redondas têm transmissão ao vivo COM Libras. "
     "São 6 momentos: abertura/Vanessa (17/08), mesa Tecnologia Delas e palestra Tessarolo (18/08), "
     "mesa Desafios da EaD e palestras Mariano e Mauro (19/08), premiação (20/08)."),
    ("Podcast nunca durante transmissão",
     "Gravação só nas janelas sem palestra/mesa no ar — mostras, sessões técnicas, almoço e pós-coffee. "
     "Janelas bloqueadas: 17/08 18h30–20h30 · 18/08 9h30–11h30 e 14h30–16h · "
     "19/08 9h30–11h30 e 14h30–17h30 · 20/08 16h."),
    ("Podcast só depois da fala",
     "O roteiro usa a transcrição da palestra, então a gravação acontece sempre DEPOIS da fala do convidado."),
    ("Termo de imagem e voz",
     "Sem termo assinado, não grava. Definir por dia quem é o responsável por colher as assinaturas "
     "(em 2024 era uma pessoa por dia) e onde os termos ficam guardados."),
    ("Todas as palestras e mesas são no Pátio",
     "Definido em 28/07/2026. Sessões técnicas e mostra de produtos seguem sem local (pedido à Coordenação)."),
]:
    r = linha(ws, r, t, "", d)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=2)
    ws.row_dimensions[r - 1].height = 44

r += 1
r = secao(ws, r, "COMO PREENCHER")
for t, d in [
    ("___", "Campo a preencher — falta definir a pessoa."),
    ("🔴", "Informação que ainda não temos (local, host, responsável)."),
    ("🟡", "Provável, aguardando confirmação."),
    ("🔒", "Transmissão no ar — janela bloqueada para podcast."),
    ("⚠️", "Conflito ou risco que precisa de decisão."),
    ("⭐", "Momento forte para foto/reels — não pode faltar registro."),
    ("Célula vazia", "Não se aplica àquela atividade."),
    ("Linha azul-clara", "Palestra ou mesa com transmissão + Libras."),
    ("Linha amarelo-clara", "Coffee, momento cultural ou celebração."),
]:
    r = linha(ws, r, t, "", d)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=2)

r += 1
r = secao(ws, r, "PENDÊNCIAS QUE TRAVAM ESTA PLANILHA")
for t, d in [
    ("Locais das sessões técnicas e da mostra de produtos",
     "Pedido à Viviane (Coordenação) em 28/07. São 3 blocos de sessões técnicas (18 e 19/08) "
     "e 2 blocos de mostra (18/08)."),
    ("Hosts dos podcasts",
     "A Coordenação do Concefor define — 1 host por episódio (5 episódios + repescagem)."),
    ("Escala nominal de transmissão, Libras, podcast e cobertura",
     "Todas as células com ___ nas 4 abas de dia. É a decisão principal a tomar nesta planilha."),
    ("Modelo de termo de autorização de imagem e voz",
     "Verificar se o Ifes tem termo institucional; imprimir cópias e definir onde ficam."),
    ("Conflito da janela do podcast do Mariano Pimentel (19/08)",
     "A palestra do Mauro Oliveira às 16h30 ocupa a janela pós-coffee prevista. Ver aba 1908."),
    ("Confirmação dos locais UAB / UnAC / NTE (17/08)",
     "UAB e UnAC provavelmente no Pátio; NTE no Auditório."),
]:
    r = linha(ws, r, t, "", d)
    ws.merge_cells(start_row=r - 1, start_column=2, end_row=r - 1, end_column=2)
    ws.row_dimensions[r - 1].height = 32

ws.freeze_panes = "A4"
ws.page_setup.orientation = "landscape"

destino = (r"c:/dev/concefor-innovation/comunicacao/planejamento/"
           r"ATUACAO DA CGTE NO CONCEFOR 2026.xlsx")
wb.save(destino)
print("salvo:", destino)
